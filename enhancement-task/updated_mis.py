from openpyxl import load_workbook
import pandas as pd

MIS_FILE = "MIS_Draft_Edited.xlsx"
SUMMARY_FILE = "tasksummary.xlsx"
OUTPUT_FILE = "MIS_Draft_Edited_Updated.xlsx"

# Read summary file
summary_df = pd.read_excel(SUMMARY_FILE)

# Build lookup
lookup = {}

for _, row in summary_df.iterrows():

    task_name = str(row["Task Name"]).strip()

    delayed = int(row["DelayedCount"])
    ontime = int(row["OnTimeCount"])

    lookup[task_name] = {
        "total_tasks": delayed + ontime,
        "delayed_count": delayed
    }

# Open MIS workbook
wb = load_workbook(MIS_FILE)
ws = wb.active

# Process column = C
# Total Tasks column = E
# Timeliness Resolution column = G

for row in range(1, ws.max_row + 1):

    process = ws[f"C{row}"].value

    if process is None:
        continue

    process = str(process).strip()

    if process in lookup:

        ws[f"E{row}"] = lookup[process]["total_tasks"]

        ws[f"G{row}"] = lookup[process]["delayed_count"]

        print(
            f"Updated {process} "
            f"-> Total={lookup[process]['total_tasks']} "
            f"Delayed={lookup[process]['delayed_count']}"
        )

wb.save(OUTPUT_FILE)

print("\nMIS UPDATED SUCCESSFULLY")
print(f"Output File: {OUTPUT_FILE}")