import os
import sys

import pandas as pd
from openpyxl import load_workbook


CONFIG_FILE = "taskchecker.json"
INPUT_TASK_FILE = "taskout1.xlsx"
SUMMARY_FILE = "tasksummary.xlsx"
MIS_FILE = "MIS_Draft_Edited.xlsx"
UPDATED_MIS_FILE = "MIS_Draft_Edited_Updated.xlsx"


def validate_files():

    if not os.path.exists(CONFIG_FILE):
        raise FileNotFoundError(
            f"Configuration file not found: {CONFIG_FILE}"
        )

    if not os.path.exists(INPUT_TASK_FILE):
        raise FileNotFoundError(
            f"Source file not found: {INPUT_TASK_FILE}"
        )

    if not os.path.exists(MIS_FILE):
        raise FileNotFoundError(
            f"MIS file not found: {MIS_FILE}"
        )


def generate_task_summary():

    print("\nGenerating Task Summary...\n")

    df = pd.read_excel(INPUT_TASK_FILE)

    if df.empty:
        raise ValueError(
            f"{INPUT_TASK_FILE} contains no data"
        )

    df.columns = df.columns.str.strip()

    required_columns = [
        "Task Name",
        "Task Date",
        "Completed On Time"
    ]

    missing_columns = (
        set(required_columns)
        - set(df.columns)
    )

    if missing_columns:
        raise ValueError(
            f"Missing columns: "
            f"{', '.join(missing_columns)}"
        )

    df = df[df["Task Name"].notna()]
    df = df[
        df["Task Name"]
        .astype(str)
        .str.strip()
        != ""
    ]

    df["Completed On Time"] = (
        df["Completed On Time"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    summary = (
        df.groupby(
            [
                "Task Name",
                "Completed On Time"
            ]
        )
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )

    if "DELAYED" not in summary.columns:
        summary["DELAYED"] = 0

    if "ON TIME" not in summary.columns:
        summary["ON TIME"] = 0

    summary = summary.rename(
        columns={
            "DELAYED": "DelayedCount",
            "ON TIME": "OnTimeCount"
        }
    )

    summary["TotalTasks"] = (
        summary["DelayedCount"]
        + summary["OnTimeCount"]
    )

    summary = summary[
        [
            "Task Name",
            "DelayedCount",
            "OnTimeCount",
            "TotalTasks"
        ]
    ]

    summary = summary.sort_values(
        "Task Name"
    )

    summary.columns.name = None

    if summary.empty:
        raise ValueError(
            "No summary data generated"
        )

    summary.to_excel(
        SUMMARY_FILE,
        index=False
    )

    print("TASK SUMMARY GENERATED\n")
    print(summary.to_string(index=False))

    return summary


def update_mis(summary):

    print("\nUpdating MIS File...\n")

    workbook = load_workbook(MIS_FILE)

    sheet = workbook.active

    task_lookup = {}

    for _, row in summary.iterrows():

        task_lookup[
            str(row["Task Name"]).strip()
        ] = {
            "total_tasks": int(
                row["TotalTasks"]
            ),
            "delayed_count": int(
                row["DelayedCount"]
            )
        }

    updated_rows = 0

    for row in range(
        1,
        sheet.max_row + 1
    ):

        task_name = sheet[
            f"C{row}"
        ].value

        if task_name is None:
            continue

        task_name = str(
            task_name
        ).strip()

        if task_name in task_lookup:

            sheet[
                f"E{row}"
            ] = task_lookup[
                task_name
            ]["total_tasks"]

            sheet[
                f"J{row}"
            ] = task_lookup[
                task_name
            ]["delayed_count"]

            updated_rows += 1

    if updated_rows == 0:
        raise ValueError(
            "No matching tasks found in MIS file"
        )

    workbook.save(
        UPDATED_MIS_FILE
    )

    print(
        f"\nMIS File Updated Successfully"
    )
    print(
        f"Rows Updated : {updated_rows}"
    )
    print(
        f"Output File : {UPDATED_MIS_FILE}"
    )


def main():

    try:

        validate_files()

        summary = generate_task_summary()

        update_mis(summary)

        print(
            "\nPROCESS COMPLETED SUCCESSFULLY"
        )

    except Exception as ex:

        print(
            f"\nERROR : {str(ex)}"
        )
        sys.exit(1)


if __name__ == "__main__":
    main()